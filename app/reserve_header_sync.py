"""Синхронизация блоков проекта на «Справка по резервам» с секцией «Утвержденные резервы» источника."""

from __future__ import annotations

import logging
from datetime import datetime, date
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _set_cell_resolve_merge(ws: Worksheet, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(rng.min_row, rng.min_col).value = value
                return
        return
    cell.value = value


def _get_cell_resolve_merge(ws: Worksheet, row: int, col: int) -> Any:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return None

RESERVE_ROLE_LABELS = frozenset({"рп", "ргп", "згд", "гд"})


def _normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").replace("\u2009", " ").strip()
    if not text:
        return ""
    return " ".join(text.lower().replace("ё", "е").split())


def _normalize_header(value: Any) -> str:
    return _normalize_key(value)


def _norm_role(cell_val: Any) -> str | None:
    nk = _normalize_header(cell_val)
    return nk if nk in RESERVE_ROLE_LABELS else None


def _cell_to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, (datetime, date)):
        return None
    t = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        pass
    if "/" in t:
        parts = t.split("/", 1)
        try:
            a, b = float(parts[0]), float(parts[1])
            if b != 0:
                return a / b
        except (ValueError, ZeroDivisionError):
            return None
    return None


def _row_matches_header_preset(sheet: Worksheet, row_idx: int, preset: dict[str, Any]) -> bool:
    headers_in_row = {_normalize_header(sheet.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, sheet.max_column + 1)}
    headers_in_row.discard("")

    for spec in preset.values():
        if not isinstance(spec, dict):
            continue
        candidates: list[str] = []
        h = spec.get("header")
        if isinstance(h, str) and h.strip():
            candidates.append(h)
        aliases = spec.get("header_aliases")
        if isinstance(aliases, list):
            candidates.extend(a for a in aliases if isinstance(a, str) and a.strip())
        if not candidates:
            continue
        norm_cand = {_normalize_header(c) for c in candidates}
        if headers_in_row.isdisjoint(norm_cand):
            return False
    return True


def _find_registry_header_row(sheet: Worksheet, preset: dict[str, Any], max_scan: int = 200) -> int | None:
    for r in range(1, min(sheet.max_row, max_scan) + 1):
        if _row_matches_header_preset(sheet, r, preset):
            return r
    return None


def _find_merge_for_substring(sheet: Worksheet, substring: str) -> Any | None:
    key = _normalize_key(substring)
    if not key:
        return None
    for rng in sheet.merged_cells.ranges:
        val = sheet.cell(rng.min_row, rng.min_col).value
        if val is None:
            continue
        if key in _normalize_key(val):
            return rng
    return None


def _find_cell_for_text(sheet: Worksheet, text: str, *, max_rows: int = 400, max_cols: int = 80) -> tuple[int, int] | None:
    target = _normalize_key(text)
    if not target:
        return None
    for r in range(1, min(sheet.max_row, max_rows) + 1):
        for c in range(1, min(sheet.max_column, max_cols) + 1):
            if _normalize_key(_get_cell_resolve_merge(sheet, r, c)) == target:
                return r, c
    return None


def _find_project_header_cells(sheet: Worksheet, *, max_rows: int = 400, max_cols: int = 80) -> list[tuple[int, int]]:
    hits: list[tuple[int, int]] = []
    target = _normalize_key("Наименование проекта")
    for r in range(1, min(sheet.max_row, max_rows) + 1):
        for c in range(1, min(sheet.max_column, max_cols) + 1):
            if _normalize_key(_get_cell_resolve_merge(sheet, r, c)) == target:
                hits.append((r, c))
    return hits


def _find_sheet_by_name_local(workbook: Workbook, sheet_name: str) -> Worksheet | None:
    target = _normalize_key(sheet_name)
    for candidate in workbook.sheetnames:
        if _normalize_key(candidate) == target:
            return workbook[candidate]
    return None


def _row_has_numeric_in_columns(sheet: Worksheet, row_idx: int, columns: list[int]) -> bool:
    for c in columns:
        v = _cell_to_float(sheet.cell(row_idx, c).value)
        if v is not None:
            return True
    return False


def _project_match_keys(value: Any) -> set[str]:
    raw = _normalize_key(value)
    if not raw:
        return set()

    keys = {raw}
    if raw.startswith("проект "):
        tail = raw.removeprefix("проект ").strip()
        if tail:
            keys.add(tail)
    elif _is_plain_number(raw):
        keys.add(f"проект {raw}")
    return keys


def _format_formula_number(value: float) -> str:
    return format(float(value), ".15g").replace(",", ".")


def _formula_term(value: Any) -> str | None:
    if isinstance(value, (int, float)):
        return _format_formula_number(float(value))

    num = _cell_to_float(value)
    if num is not None:
        return _format_formula_number(num)
    return None


def _append_formula_addend(sheet: Worksheet, row: int, col: int, addend: float) -> None:
    current = _get_cell_resolve_merge(sheet, row, col)
    new_term = _format_formula_number(addend)

    if current in (None, ""):
        next_value = f"={new_term}"
    elif isinstance(current, str) and current.strip().startswith("="):
        base = current.strip()[1:].strip()
        next_value = f"={base}+{new_term}" if base else f"={new_term}"
    else:
        base_term = _formula_term(current)
        if base_term is None:
            logger.warning("Ячейка %s%d содержит нечисловое значение %r — резерв не добавлен.", sheet.cell(row, col).column_letter, row, current)
            return
        next_value = f"={base_term}+{new_term}"

    _set_cell_resolve_merge(sheet, row, col, next_value)


def extract_approved_reserve_snapshots(
    sheet: Worksheet,
    *,
    preset: dict[str, Any],
    merge_substring: str = "Утвержденные резервы",
    project_summary_column: str = "B",
    scan_description_columns: tuple[str, str] = ("C", "D"),
) -> list[dict[str, Any]]:
    snapshots: list[dict[str, Any]] = []
    merge_rng = _find_merge_for_substring(sheet, merge_substring)
    if merge_rng is None:
        logger.debug(
            "Блок «%s» на листе %s не найден — пропуск извлечения утверждённых резервов.",
            merge_substring,
            sheet.title,
        )
        return snapshots

    header_row = int(merge_rng.max_row) + 1
    project_header = _find_cell_for_text(sheet, "Наименование проекта")
    proj_col = project_header[1] if project_header else column_index_from_string(project_summary_column.strip().upper())
    desc_cols = [column_index_from_string(a.strip().upper()) for a in scan_description_columns]

    roles_by_col: dict[int, str] = {}
    for c in range(int(merge_rng.min_col), int(merge_rng.max_col) + 1):
        label = _norm_role(sheet.cell(header_row, c).value)
        if label:
            roles_by_col[c] = label

    registry_row = _find_registry_header_row(sheet, preset)
    if registry_row is None:
        logger.warning("Не найдена строка регистра — утверждённые резервы не извлекаются (%s).", sheet.title)
        return snapshots

    first_data_row = (project_header[0] + 1) if project_header else (header_row + 2)
    for row_idx in range(first_data_row, registry_row):
        if not roles_by_col:
            break
        if not _row_has_numeric_in_columns(sheet, row_idx, list(roles_by_col.keys())):
            continue
        project_raw = _get_cell_resolve_merge(sheet, row_idx, proj_col)
        if project_raw in (None, ""):
            continue

        role_values: dict[str, float] = {}
        ok = False
        for col_idx, nk in roles_by_col.items():
            val = _cell_to_float(sheet.cell(row_idx, col_idx).value)
            if val is not None:
                role_values[nk] = val
                ok = True
        if not ok:
            continue

        description: Any = None
        for dc in desc_cols:
            v = sheet.cell(row_idx, dc).value
            if v not in (None, ""):
                description = v
                break

        snapshots.append(
            {
                "row": row_idx,
                "project_raw": project_raw,
                "roles": role_values,
                "description": description,
            }
        )

    logger.debug("Лист %s: извлечено %s проектных строк утверждённых резервов.", sheet.title, len(snapshots))
    return snapshots


def _display_project_name(project_raw: Any, source_records: list[dict[str, Any]]) -> str:
    if project_raw in (None, ""):
        return ""
    key = str(project_raw).strip()
    if not key:
        return ""

    for rec in source_records:
        rv = rec.get("project")
        if rv is None or rv == "":
            continue
        rs = str(rv).strip()
        if rs == key and not _is_plain_number(rs):
            return rs

    if _is_plain_number(key):
        return f"Проект {key}"
    return key


def _is_plain_number(s: str) -> bool:
    s = s.replace(",", ".")
    try:
        float(s)
        return True
    except ValueError:
        return s.isdigit()


def enrich_snapshots_labels(snapshots: list[dict[str, Any]], source_records: list[dict[str, Any]]) -> None:
    for snap in snapshots:
        snap["project_label"] = _display_project_name(snap.get("project_raw"), source_records)


def find_template_reserve_plan_rows(sheet: Worksheet, label_substring: str = "План по ПД с резервами") -> list[int]:
    nk = _normalize_key(label_substring)
    hits: list[int] = []
    for r in range(1, min(sheet.max_row, 400) + 1):
        for c in range(1, min(sheet.max_column, 22) + 1):
            v = sheet.cell(r, c).value
            if isinstance(v, str) and nk in _normalize_key(v):
                hits.append(r)
                break
    hits.sort()
    return hits


def split_detail_rows_by_cfo_sections(sheet: Worksheet, detail_rows: list[int]) -> list[list[int]]:
    markers: list[int] = []
    for r in range(1, min(sheet.max_row, 500) + 1):
        if _normalize_key(sheet.cell(r, 2).value) == _normalize_key("ЦФО") and _normalize_key(sheet.cell(r, 4).value) == _normalize_key(
            "Уровень резерва"
        ):
            markers.append(r)

    markers.sort()
    if not markers:
        return [detail_rows]

    segments: list[list[int]] = []
    for i, mh in enumerate(markers):
        end = markers[i + 1] if i + 1 < len(markers) else sheet.max_row + 1
        seg = sorted(d for d in detail_rows if mh < d < end)
        if seg:
            segments.append(seg)

    return segments if segments else [detail_rows]


def _find_role_value_cells_near_plan(sheet: Worksheet, plan_row: int) -> dict[str, tuple[int, int]]:
    role_cells: dict[str, tuple[int, int]] = {}
    for col in range(1, min(sheet.max_column, 80) + 1):
        role = _norm_role(_get_cell_resolve_merge(sheet, plan_row, col))
        if role:
            role_cells[role] = (plan_row + 1, col)

    if len(role_cells) < 2 and plan_row > 1:
        for col in range(1, min(sheet.max_column, 80) + 1):
            role = _norm_role(_get_cell_resolve_merge(sheet, plan_row - 1, col))
            if role:
                role_cells[role] = (plan_row + 1, col)

    return role_cells


def _find_template_project_blocks(
    sheet: Worksheet,
    *,
    plan_label_substring: str = "План по ПД с резервами",
) -> list[dict[str, Any]]:
    project_headers = _find_project_header_cells(sheet)
    plan_key = _normalize_key(plan_label_substring)
    blocks: list[dict[str, Any]] = []

    for idx, (header_row, project_col) in enumerate(project_headers):
        next_header_row = project_headers[idx + 1][0] if idx + 1 < len(project_headers) else sheet.max_row + 1
        project_row = header_row + 1
        project_value = _get_cell_resolve_merge(sheet, project_row, project_col)
        if project_value in (None, ""):
            continue

        plan_row: int | None = None
        for row_idx in range(project_row + 1, min(next_header_row, sheet.max_row + 1)):
            for col_idx in range(project_col, min(sheet.max_column, 80) + 1):
                value = _get_cell_resolve_merge(sheet, row_idx, col_idx)
                if isinstance(value, str) and plan_key in _normalize_key(value):
                    plan_row = row_idx
                    break
            if plan_row is not None:
                break

        if plan_row is None:
            continue

        blocks.append(
            {
                "project_value": project_value,
                "project_row": project_row,
                "plan_row": plan_row,
                "role_cells": _find_role_value_cells_near_plan(sheet, plan_row),
            }
        )

    return blocks


def _snapshot_key_candidates(snapshot: dict[str, Any]) -> set[str]:
    keys = _project_match_keys(snapshot.get("project_raw"))
    keys.update(_project_match_keys(snapshot.get("project_label")))
    return keys


def apply_snapshots_to_template(
    sheet: Worksheet,
    snapshots: list[dict[str, Any]],
    anchor_rows: list[int],
) -> None:
    if not snapshots or not anchor_rows:
        return

    template_blocks = _find_template_project_blocks(sheet)
    fallback_used_blocks: set[int] = set()

    for snap in snapshots:
        snap_keys = _snapshot_key_candidates(snap)
        if not snap_keys or not (snap.get("roles") or {}):
            continue

        matched_idx = None
        for idx, block in enumerate(template_blocks):
            if snap_keys & _project_match_keys(block.get("project_value")):
                matched_idx = idx
                break

        if matched_idx is None:
            for idx, block in enumerate(template_blocks):
                if idx not in fallback_used_blocks:
                    matched_idx = idx
                    fallback_used_blocks.add(idx)
                    logger.debug(
                        "Проект %r не найден в шаблоне — применяем утверждённые резервы к блоку %r по порядку.",
                        snap.get("project_raw"),
                        block.get("project_value"),
                    )
                    break

        if matched_idx is None:
            break

        block = template_blocks[matched_idx]
        role_cells = block.get("role_cells") or {}
        for role, value in (snap.get("roles") or {}).items():
            target = role_cells.get(role)
            if target is not None and value is not None:
                _append_formula_addend(sheet, target[0], target[1], float(value))


def apply_reserve_sheet_header_sync(
    template_sheet: Worksheet,
    source_books: list[tuple[str, Any]],
    source_sheet_name: str,
    *,
    preset: dict[str, Any],
    sync_cfg: dict[str, Any] | None,
    source_records: list[dict[str, Any]],
) -> None:
    if not isinstance(sync_cfg, dict) or not sync_cfg.get("enabled", False):
        return

    merge_substring = str(sync_cfg.get("approved_reserves_merge_substring") or "Утвержденные резервы")
    proj_col = str(sync_cfg.get("project_summary_column") or "B")
    plan_lab = str(sync_cfg.get("plan_row_label_substring") or "План по ПД с резервами")

    snaps: list[dict[str, Any]] = []
    for _, wb in source_books:
        if not isinstance(wb, Workbook):
            continue
        sh = _find_sheet_by_name_local(wb, source_sheet_name)
        if sh is None:
            continue
        snaps.extend(
            extract_approved_reserve_snapshots(
                sh,
                preset=preset,
                merge_substring=merge_substring,
                project_summary_column=proj_col,
            )
        )

    if not snaps:
        logger.debug("Нет данных утверждённых резервов для синхронизации шапок.")
        return

    enrich_snapshots_labels(snaps, source_records)
    anchors = find_template_reserve_plan_rows(template_sheet, plan_lab)
    if not anchors:
        logger.warning("Не найдена строка «%s» на листе %s.", plan_lab, template_sheet.title)
        return

    apply_snapshots_to_template(template_sheet, snaps, anchors)


def update_reserve_summaries_multi_block(sheet: Worksheet, detail_rows: list[int]) -> None:
    if not detail_rows:
        return

    from app.excel_service import _clean_text, _set_cell_value

    sections = split_detail_rows_by_cfo_sections(sheet, detail_rows)
    anchors = find_template_reserve_plan_rows(sheet)
    if not anchors:
        return

    for i, seg in enumerate(sections):
        if i >= len(anchors):
            break
        if not seg:
            continue
        plan_row = anchors[i]
        summary_row = plan_row - 1
        if summary_row < 2:
            continue

        start_row = min(seg)
        end_row = max(seg)

        year_label = _clean_text(sheet.cell(row=summary_row - 1, column=9).value)
        if not year_label:
            year_label = _clean_text(sheet["I3"].value)
        year_digits = "".join(ch for ch in year_label if ch.isdigit())
        target_year = year_digits[:4] if len(year_digits) >= 4 else "2025"

        _set_cell_value(sheet, f"E{summary_row}", f"=SUM(F{summary_row}:H{summary_row})")
        _set_cell_value(
            sheet,
            f"F{summary_row}",
            f'=SUMIF(G{start_row}:G{end_row},"Идеологическое изменение",F{start_row}:F{end_row})',
        )
        _set_cell_value(
            sheet,
            f"G{summary_row}",
            f'=SUMIF(G{start_row}:G{end_row},"удорожание",F{start_row}:F{end_row})',
        )
        _set_cell_value(
            sheet,
            f"H{summary_row}",
            f'=SUMIF(G{start_row}:G{end_row},"техническое",F{start_row}:F{end_row})',
        )
        _set_cell_value(
            sheet,
            f"I{summary_row}",
            f'=SUMIF(I{start_row}:I{end_row},"{target_year}",F{start_row}:F{end_row})',
        )
