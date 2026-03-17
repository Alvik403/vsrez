from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import ROOT_DIR


class ConfigError(ValueError):
    pass


DEFAULT_NEW_BLOCK_COLORS = ["E2F0D9", "DDEBF7", "FCE4D6", "FFF2CC", "E4DFEC"]


def consolidate_workbooks(
    template_bytes: bytes | None,
    template_name: str | None,
    sources: list[tuple[str, bytes]],
    config: dict[str, Any],
) -> tuple[bytes, str, dict[str, Any]]:
    resolved_template_bytes, resolved_template_name = _resolve_template_input(
        template_bytes=template_bytes,
        template_name=template_name,
        config=config,
    )

    template_suffix = Path(resolved_template_name).suffix.lower()
    keep_vba = template_suffix == ".xlsm"
    output_name = f"result_{Path(resolved_template_name).stem}{template_suffix or '.xlsx'}"

    template_book = load_workbook(
        filename=BytesIO(resolved_template_bytes),
        keep_vba=keep_vba,
    )
    template_values_book = load_workbook(
        filename=BytesIO(resolved_template_bytes),
        data_only=True,
        keep_vba=keep_vba,
    )
    source_books = [
        (
            source_name,
            load_workbook(
                filename=BytesIO(source_bytes),
                data_only=False,
                keep_vba=Path(source_name).suffix.lower() == ".xlsm",
            ),
        )
        for source_name, source_bytes in sources
    ]

    report_pages: list[dict[str, Any]] = []
    for page in config.get("pages", []):
        report_pages.append(_apply_page(template_book, template_values_book, source_books, config, page))

    _force_full_recalculation(template_book)

    output_stream = BytesIO()
    template_book.save(output_stream)
    return output_stream.getvalue(), output_name, {"pages": report_pages}


def _resolve_template_input(
    template_bytes: bytes | None,
    template_name: str | None,
    config: dict[str, Any],
) -> tuple[bytes, str]:
    if template_bytes is not None and template_name:
        return template_bytes, template_name

    default_template = config.get("default_template_path")
    if not isinstance(default_template, str) or not default_template.strip():
        raise ConfigError("Не найден шаблон: загрузите файл или укажите 'default_template_path' в конфиге.")

    template_path = (ROOT_DIR / default_template).resolve()
    if not template_path.exists():
        raise ConfigError(f"Шаблон по умолчанию не найден: {template_path}")

    return template_path.read_bytes(), template_path.name


def _apply_page(
    template_book: Workbook,
    template_values_book: Workbook,
    source_books: list[tuple[str, Workbook]],
    config: dict[str, Any],
    page: dict[str, Any],
) -> dict[str, Any]:
    page_name = str(page.get("name") or page.get("template_sheet") or "page")
    mode = str(page.get("mode") or "reserves_sheet_v1").strip().lower()

    if mode != "reserves_sheet_v1":
        raise ConfigError(f"[{page_name}] Неподдерживаемый режим страницы: '{mode}'.")

    return _apply_reserves_page(template_book, template_values_book, source_books, config, page, page_name)


def _apply_reserves_page(
    template_book: Workbook,
    template_values_book: Workbook,
    source_books: list[tuple[str, Workbook]],
    config: dict[str, Any],
    page: dict[str, Any],
    page_name: str,
) -> dict[str, Any]:
    template_sheet_name = _require_str(page, "template_sheet", page_name)
    source_sheet_name = _require_str(page, "source_sheet", page_name)
    source_preset_name = str(page.get("source_preset", "default"))
    data_start_row = int(page.get("template_start_row", 7))
    helper_sheet_name = str(page.get("helper_sheet", "Справка"))
    helper_start_row = int(page.get("helper_start_row", 301))

    if template_sheet_name not in template_book.sheetnames:
        raise ConfigError(f"[{page_name}] В шаблоне нет листа '{template_sheet_name}'.")
    if helper_sheet_name not in template_book.sheetnames:
        raise ConfigError(f"[{page_name}] В шаблоне нет листа '{helper_sheet_name}'.")

    source_preset = _get_preset(config.get("source_cols_preset"), source_preset_name, page_name)
    source_records = _collect_source_records(
        source_books=source_books,
        source_sheet_name=source_sheet_name,
        source_preset=source_preset,
        header_row=int(page.get("source_header_row", 2)),
        start_row=int(page.get("source_start_row", 3)),
        type_map=page.get("type_map"),
        page_name=page_name,
    )

    target_sheet = template_book[template_sheet_name]
    layout = _parse_existing_layout(target_sheet, data_start_row)
    _merge_source_records(layout, source_records, page)
    detail_rows = _rewrite_reserve_sheet(target_sheet, layout, page, data_start_row)
    _rewrite_helper_sheet(
        template_book=template_book,
        helper_sheet_name=helper_sheet_name,
        target_sheet_name=template_sheet_name,
        helper_start_row=helper_start_row,
        detail_rows=detail_rows,
    )
    _update_sheet_summary(target_sheet, detail_rows)
    breakdown_rows = _rewrite_breakdown_sheet(
        template_book=template_book,
        template_values_book=template_values_book,
        page=page,
        source_records=source_records,
    )
    cfo_rows = _rewrite_cfo_sheet(
        template_book=template_book,
        template_values_book=template_values_book,
        page=page,
        source_records=source_records,
    )

    return {
        "name": page_name,
        "template_sheet": template_sheet_name,
        "source_sheet": source_sheet_name,
        "source_rows_loaded": len(source_records),
        "detail_rows_written": len(detail_rows),
        "cfo_blocks_total": len(layout["blocks"]),
        "new_cfo_blocks": sum(1 for block in layout["blocks"] if block.get("is_new")),
        "breakdown_rows_written": breakdown_rows,
        "cfo_sheet_rows_written": cfo_rows,
    }


def _collect_source_records(
    source_books: list[tuple[str, Workbook]],
    source_sheet_name: str,
    source_preset: dict[str, Any],
    header_row: int,
    start_row: int,
    type_map: Any,
    page_name: str,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    for source_name, workbook in source_books:
        sheet = _find_sheet_by_name(workbook, source_sheet_name)
        if sheet is None:
            raise ConfigError(
                f"[{page_name}] В исходном файле '{source_name}' нет листа '{source_sheet_name}'."
            )

        resolved_header_row = header_row if header_row > 0 else _find_header_row(sheet, source_preset, page_name, source_name)
        resolved_start_row = max(start_row, resolved_header_row + 1) if start_row > 0 else resolved_header_row + 1
        columns = {
            field_name: _resolve_preset_column(
                sheet,
                source_preset,
                field_name,
                resolved_header_row,
                f"[{page_name}] исходник {source_name}",
            )
            for field_name in source_preset
        }

        for row_idx in range(resolved_start_row, sheet.max_row + 1):
            cfo = _clean_text(sheet.cell(row=row_idx, column=columns["cfo"]).value)
            level = _clean_text(sheet.cell(row=row_idx, column=columns["level"]).value)
            amount = sheet.cell(row=row_idx, column=columns["amount"]).value
            if not cfo or not level or amount in (None, ""):
                continue

            records.append(
                {
                    "source_name": source_name,
                    "project": _clean_text(sheet.cell(row=row_idx, column=columns["project"]).value),
                    "cfo": cfo,
                    "level": level,
                    "reason": _clean_text(sheet.cell(row=row_idx, column=columns["reason"]).value),
                    "comment": _clean_text(sheet.cell(row=row_idx, column=columns["comment"]).value),
                    "type": _normalize_type(
                        sheet.cell(row=row_idx, column=columns["type"]).value,
                        type_map,
                    ),
                    "amount": _to_number(amount),
                    "date": _extract_year(sheet.cell(row=row_idx, column=columns["date"]).value),
                    "work_value": None,
                    "is_new": True,
                }
            )

    return records


def _parse_existing_layout(sheet: Worksheet, start_row: int) -> dict[str, Any]:
    blocks: list[dict[str, Any]] = []
    blank_template = None
    current_block: dict[str, Any] | None = None
    current_level: dict[str, Any] | None = None

    for row_idx in range(start_row, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(2, 11)]
        if all(value in (None, "") for value in row_values):
            blank_template = blank_template or _snapshot_row(sheet, row_idx)
            continue

        cfo = _clean_text(sheet[f"B{row_idx}"].value)
        level = _clean_text(sheet[f"D{row_idx}"].value)
        row_type = _classify_sheet_row(sheet, row_idx)

        if row_type == "cfo_total":
            current_block = {
                "name": cfo,
                "levels": [],
                "is_new": False,
                "templates": {"cfo_total": _snapshot_row(sheet, row_idx)},
            }
            blocks.append(current_block)
            current_level = None
            continue

        if row_type == "level_total":
            if current_block is None:
                continue
            current_level = {
                "name": level,
                "details": [],
                "template_total": _snapshot_row(sheet, row_idx),
            }
            current_block["levels"].append(current_level)
            current_block["templates"].setdefault("level_total", current_level["template_total"])
            continue

        if row_type == "detail" and current_block is not None and current_level is not None:
            record = {
                "project": _clean_text(sheet[f"J{row_idx}"].value),
                "cfo": cfo or current_block["name"],
                "level": level or current_level["name"],
                "reason": _clean_text(sheet[f"E{row_idx}"].value) if not _is_formula(sheet[f"E{row_idx}"].value) else "",
                "comment": _clean_text(sheet[f"H{row_idx}"].value),
                "type": _clean_text(sheet[f"G{row_idx}"].value),
                "amount": _to_number(sheet[f"F{row_idx}"].value),
                "date": _extract_year(sheet[f"I{row_idx}"].value),
                "work_value": sheet[f"E{row_idx}"].value,
                "is_new": False,
            }
            current_level["details"].append(record)
            current_block["templates"].setdefault("detail", _snapshot_row(sheet, row_idx))
            current_block["templates"]["detail_end"] = _snapshot_row(sheet, row_idx)

    if not blocks:
        raise ConfigError("Не удалось распознать структуру листа 'Справка по резервам '.")

    for block in blocks:
        if "level_total" not in block["templates"]:
            raise ConfigError(f"Блок ЦФО '{block['name']}' не содержит строки уровня резерва.")
        if "detail" not in block["templates"]:
            block["templates"]["detail"] = block["templates"]["level_total"]
            block["templates"]["detail_end"] = block["templates"]["level_total"]

    if blank_template is None:
        blank_template = _snapshot_row(sheet, sheet.max_row)

    return {"blocks": blocks, "blank_template": blank_template}


def _merge_source_records(layout: dict[str, Any], source_records: list[dict[str, Any]], page: dict[str, Any]) -> None:
    blocks = layout["blocks"]
    color_palette = page.get("new_block_fill_colors") or DEFAULT_NEW_BLOCK_COLORS
    reference_block = blocks[-1]
    new_block_index = 0

    for record in source_records:
        block = _find_block(blocks, record["cfo"])
        if block is None:
            block = {
                "name": record["cfo"],
                "levels": [],
                "is_new": True,
                "fill_color": color_palette[new_block_index % len(color_palette)],
                "templates": _clone_templates(reference_block["templates"]),
            }
            blocks.append(block)
            new_block_index += 1

        level = _find_level(block["levels"], record["level"])
        if level is None:
            level = {
                "name": record["level"],
                "details": [],
                "template_total": copy(block["templates"]["level_total"]),
            }
            block["levels"].append(level)

        level["details"].append(record)


def _rewrite_reserve_sheet(
    sheet: Worksheet,
    layout: dict[str, Any],
    page: dict[str, Any],
    start_row: int,
) -> list[int]:
    detail_rows: list[int] = []
    row_idx = start_row
    first_project = _pick_first_project(layout["blocks"])

    if first_project:
        sheet["B4"] = first_project

    for block in layout["blocks"]:
        fill_color = block.get("fill_color") if block.get("is_new") else None

        _apply_row_template(sheet, row_idx, block["templates"]["cfo_total"], fill_color)
        cfo_total_row = row_idx
        sheet[f"B{row_idx}"] = block["name"]
        _clear_cells(sheet, row_idx, ["C", "D", "E", "G", "H", "I", "J"])
        row_idx += 1

        level_total_rows: list[int] = []
        for level_index, level in enumerate(block["levels"]):
            _apply_row_template(sheet, row_idx, level["template_total"], fill_color)
            level_total_row = row_idx
            level_total_rows.append(level_total_row)
            sheet[f"B{row_idx}"] = block["name"]
            sheet[f"D{row_idx}"] = level["name"]
            _clear_cells(sheet, row_idx, ["C", "E", "G", "H", "I", "J"])
            row_idx += 1

            level_detail_rows: list[int] = []
            is_last_level = level_index == len(block["levels"]) - 1
            for detail_index, detail in enumerate(level["details"]):
                is_last_detail_in_block = is_last_level and detail_index == len(level["details"]) - 1
                template_key = "detail_end" if is_last_detail_in_block else "detail"
                _apply_row_template(sheet, row_idx, block["templates"][template_key], fill_color)
                _write_detail_row(sheet, row_idx, detail, block["name"], level["name"])
                detail_rows.append(row_idx)
                level_detail_rows.append(row_idx)
                row_idx += 1

            sheet[f"F{level_total_row}"] = _build_sum_formula(level_detail_rows)

        sheet[f"F{cfo_total_row}"] = _build_addition_formula(level_total_rows)

    blank_template = layout["blank_template"]
    for _ in range(3):
        _apply_row_template(sheet, row_idx, blank_template, None)
        _clear_cells(sheet, row_idx, [get_column_letter(col) for col in range(1, 11)])
        row_idx += 1

    for clear_row in range(row_idx, sheet.max_row + 1):
        _clear_cells(sheet, clear_row, [get_column_letter(col) for col in range(1, 11)])

    return detail_rows


def _rewrite_helper_sheet(
    template_book: Workbook,
    helper_sheet_name: str,
    target_sheet_name: str,
    helper_start_row: int,
    detail_rows: list[int],
) -> None:
    helper_sheet = template_book[helper_sheet_name]
    helper_template = _snapshot_row(helper_sheet, helper_start_row)
    max_existing = helper_sheet.max_row

    for index, source_row in enumerate(detail_rows):
        row_idx = helper_start_row + index
        _apply_row_template(helper_sheet, row_idx, helper_template, None)
        helper_sheet[f"B{row_idx}"] = f"='{target_sheet_name}'!D{source_row}"
        helper_sheet[f"C{row_idx}"] = f"='{target_sheet_name}'!E{source_row}"
        helper_sheet[f"D{row_idx}"] = f"='{target_sheet_name}'!F{source_row}"
        helper_sheet[f"E{row_idx}"] = f"='{target_sheet_name}'!G{source_row}"
        helper_sheet[f"F{row_idx}"] = f"='{target_sheet_name}'!H{source_row}"
        helper_sheet[f"G{row_idx}"] = f"='{target_sheet_name}'!I{source_row}"
        helper_sheet[f"H{row_idx}"] = f"='{target_sheet_name}'!J{source_row}"
        helper_sheet[f"I{row_idx}"] = f"='{target_sheet_name}'!B{source_row}"

    for row_idx in range(helper_start_row + len(detail_rows), max_existing + 1):
        _clear_cells(helper_sheet, row_idx, ["B", "C", "D", "E", "F", "G", "H", "I"])


def _update_sheet_summary(sheet: Worksheet, detail_rows: list[int]) -> None:
    if not detail_rows:
        return

    start_row = min(detail_rows)
    end_row = max(detail_rows)
    year_label = _clean_text(sheet["I3"].value)
    year_digits = "".join(ch for ch in year_label if ch.isdigit())
    target_year = year_digits or "2025"

    sheet["E4"] = "=SUM(F4:H4)"
    sheet["F4"] = f'=SUMIF(G{start_row}:G{end_row},"Идеологическое изменение",F{start_row}:F{end_row})'
    sheet["G4"] = f'=SUMIF(G{start_row}:G{end_row},"удорожание",F{start_row}:F{end_row})'
    sheet["H4"] = f'=SUMIF(G{start_row}:G{end_row},"техническое",F{start_row}:F{end_row})'
    sheet["I4"] = f'=SUMIF(I{start_row}:I{end_row},"{target_year}",F{start_row}:F{end_row})'


def _rewrite_breakdown_sheet(
    template_book: Workbook,
    template_values_book: Workbook,
    page: dict[str, Any],
    source_records: list[dict[str, Any]],
) -> int:
    sheet_name = str(page.get("breakdown_sheet", "Виды вскрытия 2"))
    if sheet_name not in template_book.sheetnames:
        return 0

    sheet = template_book[sheet_name]
    value_sheet = template_values_book[sheet_name] if sheet_name in template_values_book.sheetnames else sheet
    start_row = int(page.get("breakdown_start_row", 6))
    type_template = _snapshot_row(sheet, int(page.get("breakdown_type_template_row", 6)))
    comment_template = _snapshot_row(sheet, int(page.get("breakdown_comment_template_row", 7)))
    reason_template = _snapshot_row(sheet, int(page.get("breakdown_reason_template_row", 8)))
    total_template = _snapshot_row(sheet, int(page.get("breakdown_total_template_row", 32)))
    type_order = page.get("type_order") if isinstance(page.get("type_order"), list) else []
    grouped = _parse_existing_breakdown_sheet(
        sheet=sheet,
        value_sheet=value_sheet,
        start_row=start_row,
        total_label=str(page.get("breakdown_total_label", "Общий итог")),
    )
    _merge_breakdown_source_records(
        grouped=grouped,
        source_records=source_records,
        type_template=type_template,
        comment_template=comment_template,
        reason_template=reason_template,
        type_order=type_order,
    )

    row_idx = start_row
    max_existing = sheet.max_row
    for type_entry in grouped:
        _apply_row_template(sheet, row_idx, type_entry.get("template", type_template), None)
        _set_cell_value(sheet, f"A{row_idx}", type_entry.get("output_a", type_entry["name"]))
        type_row_idx = row_idx
        type_comment_rows: list[int] = []
        _set_cell_value(sheet, f"B{row_idx}", None)
        _set_cell_value(sheet, f"C{row_idx}", type_entry.get("output_c"))
        _set_cell_value(sheet, f"D{row_idx}", type_entry.get("output_d"))
        _set_row_grouping(
            sheet,
            row_idx,
            outline_level=0,
            hidden=False,
            collapsed=bool(type_entry["comments"]),
        )
        row_idx += 1

        for comment_entry in type_entry["comments"]:
            _apply_row_template(sheet, row_idx, comment_entry.get("template", comment_template), None)
            _set_cell_value(sheet, f"A{row_idx}", comment_entry.get("output_a", comment_entry["name"]))
            comment_row_idx = row_idx
            comment_reason_rows: list[int] = []
            type_comment_rows.append(comment_row_idx)
            _set_cell_value(sheet, f"B{row_idx}", None)
            _set_cell_value(sheet, f"C{row_idx}", comment_entry.get("output_c"))
            _set_cell_value(sheet, f"D{row_idx}", comment_entry.get("output_d"))
            _set_row_grouping(
                sheet,
                row_idx,
                outline_level=1,
                hidden=True,
                collapsed=bool(comment_entry["reasons"]),
            )
            row_idx += 1

            for reason_entry in comment_entry["reasons"]:
                _apply_row_template(sheet, row_idx, reason_entry.get("template", reason_template), None)
                _set_cell_value(sheet, f"A{row_idx}", reason_entry.get("output_a", reason_entry["name"]))
                if reason_entry.get("is_existing"):
                    _set_cell_value(sheet, f"B{row_idx}", reason_entry.get("output_b"))
                    _set_cell_value(sheet, f"C{row_idx}", reason_entry.get("output_c"))
                    _set_cell_value(sheet, f"D{row_idx}", reason_entry.get("output_d"))
                else:
                    _set_cell_value(sheet, f"B{row_idx}", reason_entry["amount"])
                    _set_cell_value(sheet, f"C{row_idx}", _single_value_or_blank(reason_entry["projects"]))
                    _set_cell_value(sheet, f"D{row_idx}", _single_value_or_blank(reason_entry["cfos"]))
                comment_reason_rows.append(row_idx)
                _set_row_grouping(
                    sheet,
                    row_idx,
                    outline_level=2,
                    hidden=True,
                    collapsed=False,
                )
                row_idx += 1

            _set_cell_value(
                sheet,
                f"B{comment_row_idx}",
                _build_sum_formula_for_column(comment_reason_rows, "B"),
            )

        _set_cell_value(
            sheet,
            f"B{type_row_idx}",
            _build_addition_formula_for_column(type_comment_rows, "B"),
        )

    _apply_row_template(sheet, row_idx, total_template, None)
    _set_cell_value(sheet, f"A{row_idx}", str(page.get("breakdown_total_label", "Общий итог")))
    type_rows = [start_row + idx for idx, entry in enumerate(_flatten_breakdown_type_rows(grouped)) if entry == "type"]
    _set_cell_value(sheet, f"B{row_idx}", _build_addition_formula_for_column(type_rows, "B"))
    _set_cell_value(sheet, f"C{row_idx}", None)
    _set_cell_value(sheet, f"D{row_idx}", None)
    _set_row_grouping(
        sheet,
        row_idx,
        outline_level=0,
        hidden=False,
        collapsed=False,
    )
    row_idx += 1

    for clear_row in range(row_idx, max_existing + 1):
        _clear_cells(sheet, clear_row, ["A", "B", "C", "D"])
        _set_row_grouping(
            sheet,
            clear_row,
            outline_level=0,
            hidden=False,
            collapsed=False,
        )

    return row_idx - start_row


def _rewrite_cfo_sheet(
    template_book: Workbook,
    template_values_book: Workbook,
    page: dict[str, Any],
    source_records: list[dict[str, Any]],
) -> int:
    sheet_name = str(page.get("cfo_sheet", "Справка по ЦФО"))
    if sheet_name not in template_book.sheetnames:
        return 0

    sheet = template_book[sheet_name]
    value_sheet = template_values_book[sheet_name] if sheet_name in template_values_book.sheetnames else sheet
    style_sheet_name = str(page.get("cfo_style_source_sheet", sheet_name))
    style_sheet = template_book[style_sheet_name] if style_sheet_name in template_book.sheetnames else sheet
    start_row = int(page.get("cfo_start_row", 4))
    cfo_template = _limit_template_columns(_snapshot_row(style_sheet, int(page.get("cfo_group_template_row", 4))), 3)
    comment_template = _limit_template_columns(_snapshot_row(style_sheet, int(page.get("cfo_comment_template_row", 5))), 3)
    reason_template = _limit_template_columns(_snapshot_row(style_sheet, int(page.get("cfo_reason_template_row", 6))), 3)
    total_template = _limit_template_columns(_snapshot_row(style_sheet, int(page.get("cfo_total_template_row", 519))), 3)
    grouped = _parse_existing_cfo_sheet(
        sheet=sheet,
        value_sheet=value_sheet,
        start_row=start_row,
        total_label=str(page.get("cfo_total_label", "Общий итог")),
    )
    _merge_cfo_source_records(
        grouped=grouped,
        source_records=source_records,
        cfo_template=cfo_template,
        comment_template=comment_template,
        reason_template=reason_template,
    )

    row_idx = start_row
    max_existing = sheet.max_row
    for cfo_entry in grouped:
        _apply_row_template(sheet, row_idx, cfo_template, None)
        _set_cell_value(sheet, f"A{row_idx}", cfo_entry.get("output_a", f"Вскрытие {cfo_entry['name']}"))
        cfo_row_idx = row_idx
        cfo_comment_rows: list[int] = []
        _set_cell_value(sheet, f"B{row_idx}", None)
        _set_cell_value(sheet, f"C{row_idx}", cfo_entry.get("output_c"))
        _clear_cells(sheet, row_idx, ["D"])
        _set_row_grouping(sheet, row_idx, outline_level=0, hidden=False, collapsed=bool(cfo_entry["comments"]))
        _format_cfo_row(sheet, row_idx, min_height=cfo_template.get("height"))
        row_idx += 1

        for comment_entry in cfo_entry["comments"]:
            _apply_row_template(sheet, row_idx, comment_template, None)
            _set_cell_value(sheet, f"A{row_idx}", comment_entry.get("output_a", comment_entry["name"]))
            comment_row_idx = row_idx
            comment_reason_rows: list[int] = []
            cfo_comment_rows.append(comment_row_idx)
            _set_cell_value(sheet, f"B{row_idx}", None)
            _set_cell_value(sheet, f"C{row_idx}", comment_entry.get("output_c"))
            _clear_cells(sheet, row_idx, ["D"])
            _set_row_grouping(sheet, row_idx, outline_level=1, hidden=True, collapsed=bool(comment_entry["reasons"]))
            _format_cfo_row(sheet, row_idx, min_height=comment_template.get("height"))
            row_idx += 1

            for reason_entry in comment_entry["reasons"]:
                _apply_row_template(sheet, row_idx, reason_template, None)
                _set_cell_value(sheet, f"A{row_idx}", reason_entry.get("output_a", reason_entry["name"]))
                if reason_entry.get("is_existing"):
                    _set_cell_value(sheet, f"B{row_idx}", reason_entry.get("output_b"))
                    output_c = reason_entry.get("output_c")
                    if _is_formula(output_c) and "VLOOKUP(" in str(output_c).upper():
                        output_c = f"=VLOOKUP(A{row_idx},Справка!F:H,3,0)"
                    _set_cell_value(sheet, f"C{row_idx}", output_c)
                    _set_cell_value(sheet, f"D{row_idx}", reason_entry.get("output_d"))
                else:
                    _set_cell_value(sheet, f"B{row_idx}", reason_entry["amount"])
                    _set_cell_value(sheet, f"C{row_idx}", _project_value_or_blank(reason_entry["projects"]))
                _clear_cells(sheet, row_idx, ["D"])
                comment_reason_rows.append(row_idx)
                _set_row_grouping(sheet, row_idx, outline_level=2, hidden=True, collapsed=False)
                _format_cfo_row(sheet, row_idx, min_height=reason_template.get("height"))
                row_idx += 1

            _set_cell_value(sheet, f"B{comment_row_idx}", _build_sum_formula_for_column(comment_reason_rows, "B"))

        _set_cell_value(sheet, f"B{cfo_row_idx}", _build_addition_formula_for_column(cfo_comment_rows, "B"))

    _apply_row_template(sheet, row_idx, total_template, None)
    _set_cell_value(sheet, f"A{row_idx}", str(page.get("cfo_total_label", "Общий итог")))
    cfo_rows = [start_row + idx for idx, entry in enumerate(_flatten_breakdown_type_rows(grouped)) if entry == "type"]
    _set_cell_value(sheet, f"B{row_idx}", _build_addition_formula_for_column(cfo_rows, "B"))
    _set_cell_value(sheet, f"C{row_idx}", None)
    _clear_cells(sheet, row_idx, ["D"])
    _set_row_grouping(sheet, row_idx, outline_level=0, hidden=False, collapsed=False)
    _format_cfo_row(sheet, row_idx, min_height=total_template.get("height"))
    row_idx += 1

    for clear_row in range(row_idx, max_existing + 1):
        _clear_cells(sheet, clear_row, ["A", "B", "C", "D"])
        _set_row_grouping(sheet, clear_row, outline_level=0, hidden=False, collapsed=False)

    return row_idx - start_row


def _parse_existing_breakdown_sheet(
    sheet: Worksheet,
    value_sheet: Worksheet,
    start_row: int,
    total_label: str,
) -> list[dict[str, Any]]:
    grouped: list[dict[str, Any]] = []
    current_type: dict[str, Any] | None = None
    current_comment: dict[str, Any] | None = None
    total_key = _normalize_key(total_label)

    for row_idx in range(start_row, sheet.max_row + 1):
        label = _clean_text(value_sheet.cell(row=row_idx, column=1).value)
        if not label:
            continue
        if _normalize_key(label) == total_key:
            break

        row_dimension = sheet.row_dimensions[row_idx]
        outline_level = int(row_dimension.outlineLevel or 0)

        if outline_level == 0:
            current_type = {
                "name": label,
                "comments": [],
                "template": _snapshot_row(sheet, row_idx),
                "is_existing": True,
                "output_a": sheet.cell(row=row_idx, column=1).value,
                "output_c": sheet.cell(row=row_idx, column=3).value,
                "output_d": sheet.cell(row=row_idx, column=4).value,
            }
            grouped.append(current_type)
            current_comment = None
            continue

        if outline_level == 1 and current_type is not None:
            current_comment = {
                "name": label,
                "reasons": [],
                "template": _snapshot_row(sheet, row_idx),
                "is_existing": True,
                "output_a": sheet.cell(row=row_idx, column=1).value,
                "output_c": sheet.cell(row=row_idx, column=3).value,
                "output_d": sheet.cell(row=row_idx, column=4).value,
            }
            current_type["comments"].append(current_comment)
            continue

        if outline_level >= 2 and current_type is not None and current_comment is not None:
            current_comment["reasons"].append(
                {
                    "name": label,
                    "template": _snapshot_row(sheet, row_idx),
                    "is_existing": True,
                    "output_a": sheet.cell(row=row_idx, column=1).value,
                    "output_b": sheet.cell(row=row_idx, column=2).value,
                    "output_c": sheet.cell(row=row_idx, column=3).value,
                    "output_d": sheet.cell(row=row_idx, column=4).value,
                    "amount": 0.0,
                    "projects": [],
                    "cfos": [],
                }
            )

    return grouped


def _parse_existing_cfo_sheet(
    sheet: Worksheet,
    value_sheet: Worksheet,
    start_row: int,
    total_label: str,
) -> list[dict[str, Any]]:
    grouped: list[dict[str, Any]] = []
    current_cfo: dict[str, Any] | None = None
    current_comment: dict[str, Any] | None = None
    total_key = _normalize_key(total_label)

    for row_idx in range(start_row, sheet.max_row + 1):
        label = _clean_text(value_sheet.cell(row=row_idx, column=1).value)
        if not label:
            continue
        if _normalize_key(label) == total_key:
            break

        outline_level = int(sheet.row_dimensions[row_idx].outlineLevel or 0)

        if outline_level == 0:
            cfo_name = label.removeprefix("Вскрытие ").strip() if label.startswith("Вскрытие ") else label
            current_cfo = {
                "name": cfo_name,
                "comments": [],
                "template": _snapshot_row(sheet, row_idx),
                "is_existing": True,
                "output_a": sheet.cell(row=row_idx, column=1).value,
                "output_c": sheet.cell(row=row_idx, column=3).value,
                "output_d": sheet.cell(row=row_idx, column=4).value,
            }
            grouped.append(current_cfo)
            current_comment = None
            continue

        if outline_level == 1 and current_cfo is not None:
            current_comment = {
                "name": label,
                "reasons": [],
                "template": _snapshot_row(sheet, row_idx),
                "is_existing": True,
                "output_a": sheet.cell(row=row_idx, column=1).value,
                "output_c": sheet.cell(row=row_idx, column=3).value,
                "output_d": sheet.cell(row=row_idx, column=4).value,
            }
            current_cfo["comments"].append(current_comment)
            continue

        if outline_level >= 2 and current_cfo is not None and current_comment is not None:
            current_comment["reasons"].append(
                {
                    "name": label,
                    "template": _snapshot_row(sheet, row_idx),
                    "is_existing": True,
                    "output_a": sheet.cell(row=row_idx, column=1).value,
                    "output_b": sheet.cell(row=row_idx, column=2).value,
                    "output_c": sheet.cell(row=row_idx, column=3).value,
                    "output_d": sheet.cell(row=row_idx, column=4).value,
                    "amount": 0.0,
                    "projects": [],
                    "cfos": [],
                }
            )

    return grouped


def _merge_breakdown_source_records(
    grouped: list[dict[str, Any]],
    source_records: list[dict[str, Any]],
    type_template: dict[str, Any],
    comment_template: dict[str, Any],
    reason_template: dict[str, Any],
    type_order: list[Any],
) -> None:
    type_index = {_normalize_key(entry["name"]): entry for entry in grouped}

    def make_type(name: str) -> dict[str, Any]:
        return {
            "name": name,
            "comments": [],
            "template": type_template,
            "is_existing": False,
        }

    def make_comment(name: str) -> dict[str, Any]:
        return {
            "name": name,
            "reasons": [],
            "template": comment_template,
            "is_existing": False,
        }

    def make_reason(record: dict[str, Any], name: str) -> dict[str, Any]:
        return {
            "name": name,
            "template": reason_template,
            "is_existing": False,
            "amount_value": None,
            "project_value": None,
            "cfo_value": None,
            "amount": _to_number(record.get("amount")),
            "projects": [_clean_text(record.get("project"))] if _clean_text(record.get("project")) else [],
            "cfos": [_clean_text(record.get("cfo"))] if _clean_text(record.get("cfo")) else [],
        }

    for record in source_records:
        type_name = _clean_text(record.get("type")) or "Без типа вскрытия"
        comment_name = _clean_text(record.get("comment")) or "Без комментария экономиста"
        reason_name = _clean_text(record.get("reason")) or "Без причины"

        type_key = _normalize_key(type_name)
        type_entry = type_index.get(type_key)
        if type_entry is None:
            type_entry = make_type(type_name)
            grouped.append(type_entry)
            type_index[type_key] = type_entry

        comment_entry = next(
            (entry for entry in type_entry["comments"] if _normalize_key(entry["name"]) == _normalize_key(comment_name)),
            None,
        )
        if comment_entry is None:
            comment_entry = make_comment(comment_name)
            type_entry["comments"].append(comment_entry)

        reason_entry = next(
            (entry for entry in comment_entry["reasons"] if _normalize_key(entry["name"]) == _normalize_key(reason_name)),
            None,
        )
        if reason_entry is None:
            comment_entry["reasons"].append(make_reason(record, reason_name))
            continue

        if not reason_entry.get("is_existing"):
            reason_entry["amount"] += _to_number(record.get("amount"))
            _append_unique(reason_entry["projects"], _clean_text(record.get("project")))
            _append_unique(reason_entry["cfos"], _clean_text(record.get("cfo")))


def _merge_cfo_source_records(
    grouped: list[dict[str, Any]],
    source_records: list[dict[str, Any]],
    cfo_template: dict[str, Any],
    comment_template: dict[str, Any],
    reason_template: dict[str, Any],
) -> None:
    cfo_index = {_normalize_key(entry["name"]): entry for entry in grouped}

    def make_cfo(name: str) -> dict[str, Any]:
        return {
            "name": name,
            "comments": [],
            "template": _clone_template(cfo_template),
            "is_existing": False,
            "output_a": f"Вскрытие {name}",
        }

    def make_comment(name: str, template: dict[str, Any]) -> dict[str, Any]:
        return {
            "name": name,
            "reasons": [],
            "template": _clone_template(template),
            "is_existing": False,
        }

    def make_reason(record: dict[str, Any], name: str, template: dict[str, Any]) -> dict[str, Any]:
        project = _clean_text(record.get("project"))
        cfo = _clean_text(record.get("cfo"))
        return {
            "name": name,
            "template": _clone_template(template),
            "is_existing": False,
            "output_b": None,
            "output_c": None,
            "output_d": None,
            "amount": _to_number(record.get("amount")),
            "projects": [project] if project else [],
            "cfos": [cfo] if cfo else [],
        }

    for record in source_records:
        cfo_name = _clean_text(record.get("cfo")) or "Без ЦФО"
        comment_name = _clean_text(record.get("comment")) or "Без комментария экономиста"
        reason_name = _clean_text(record.get("reason")) or "Без причины"

        cfo_key = _normalize_key(cfo_name)
        cfo_entry = cfo_index.get(cfo_key)
        if cfo_entry is None:
            cfo_entry = make_cfo(cfo_name)
            grouped.append(cfo_entry)
            cfo_index[cfo_key] = cfo_entry

        comment_entry = next(
            (entry for entry in cfo_entry["comments"] if _normalize_key(entry["name"]) == _normalize_key(comment_name)),
            None,
        )
        if comment_entry is None:
            comment_entry = make_comment(comment_name, comment_template)
            cfo_entry["comments"].append(comment_entry)

        reason_entry = next(
            (entry for entry in comment_entry["reasons"] if _normalize_key(entry["name"]) == _normalize_key(reason_name)),
            None,
        )
        if reason_entry is None:
            comment_entry["reasons"].append(make_reason(record, reason_name, reason_template))
            continue

        if not reason_entry.get("is_existing"):
            reason_entry["amount"] += _to_number(record.get("amount"))
            _append_unique(reason_entry["projects"], _clean_text(record.get("project")))
            _append_unique(reason_entry["cfos"], _clean_text(record.get("cfo")))


def _flatten_breakdown_type_rows(grouped: list[dict[str, Any]]) -> list[str]:
    rows: list[str] = []
    for type_entry in grouped:
        rows.append("type")
        for comment_entry in type_entry["comments"]:
            rows.append("comment")
            for _reason_entry in comment_entry["reasons"]:
                rows.append("reason")
    return rows


def _write_detail_row(sheet: Worksheet, row_idx: int, detail: dict[str, Any], cfo: str, level: str) -> None:
    normalized_type = _normalize_key(detail.get("type"))
    work_value = detail.get("work_value")

    if _is_formula(work_value):
        work_value = _build_work_formula(row_idx)
    elif not work_value:
        if normalized_type == "техническое":
            work_value = detail.get("reason") or detail.get("comment")
        else:
            work_value = _build_work_formula(row_idx)

    sheet[f"B{row_idx}"] = cfo
    sheet[f"D{row_idx}"] = level
    sheet[f"E{row_idx}"] = work_value
    sheet[f"F{row_idx}"] = detail.get("amount", 0)
    sheet[f"G{row_idx}"] = detail.get("type")
    sheet[f"H{row_idx}"] = detail.get("comment")
    sheet[f"I{row_idx}"] = detail.get("date")
    sheet[f"J{row_idx}"] = detail.get("project")


def _build_work_formula(row_idx: int) -> str:
    return (
        f'=IF(G{row_idx}="Удорожание",'
        f"INDEX('статьи для сбора Удорожание'!A:C,"
        f"MATCH('Справка по резервам '!H{row_idx},'статьи для сбора Удорожание'!A:A,0),3),"
        f"INDEX('статьи для сбора Идеологическое'!A:C,"
        f"MATCH('Справка по резервам '!H{row_idx},'статьи для сбора Идеологическое'!A:A,0),3))"
    )


def _build_sum_formula(rows: list[int]) -> str:
    return _build_sum_formula_for_column(rows, "F")


def _build_sum_formula_for_column(rows: list[int], column: str) -> str:
    if not rows:
        return "=0"
    if len(rows) == 1:
        return f"={column}{rows[0]}"
    return f"=SUM({column}{rows[0]}:{column}{rows[-1]})"


def _build_addition_formula(rows: list[int]) -> str:
    return _build_addition_formula_for_column(rows, "F")


def _build_addition_formula_for_column(rows: list[int], column: str) -> str:
    if not rows:
        return "=0"
    if len(rows) == 1:
        return f"={column}{rows[0]}"
    return "=" + "+".join(f"{column}{row}" for row in rows)


def _snapshot_row(sheet: Worksheet, row_idx: int) -> dict[str, Any]:
    styles: dict[int, Any] = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        styles[col_idx] = copy(cell._style)
    merges: list[tuple[int, int]] = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == row_idx and merged_range.max_row == row_idx:
            merges.append((merged_range.min_col, merged_range.max_col))
    return {
        "height": sheet.row_dimensions[row_idx].height,
        "styles": styles,
        "merges": merges,
    }


def _apply_row_template(
    sheet: Worksheet,
    row_idx: int,
    template: dict[str, Any],
    fill_color: str | None,
) -> None:
    _reset_row_merges(sheet, row_idx)

    if template.get("height") is not None:
        sheet.row_dimensions[row_idx].height = template["height"]

    for col_idx, style in template["styles"].items():
        cell = sheet.cell(row=row_idx, column=col_idx)
        if isinstance(sheet[f"{get_column_letter(col_idx)}{row_idx}"], MergedCell):
            continue
        cell._style = copy(style)

        if fill_color and 2 <= col_idx <= 10 and cell.fill.fill_type == "solid":
            cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)

    for min_col, max_col in template.get("merges", []):
        sheet.merge_cells(
            start_row=row_idx,
            start_column=min_col,
            end_row=row_idx,
            end_column=max_col,
        )


def _clear_cells(sheet: Worksheet, row_idx: int, columns: list[str]) -> None:
    for column in columns:
        _set_cell_value(sheet, f"{column}{row_idx}", None)


def _classify_sheet_row(sheet: Worksheet, row_idx: int) -> str:
    cfo = _clean_text(sheet[f"B{row_idx}"].value)
    level = _clean_text(sheet[f"D{row_idx}"].value)
    kind = _clean_text(sheet[f"G{row_idx}"].value)
    comment = _clean_text(sheet[f"H{row_idx}"].value)
    project = _clean_text(sheet[f"J{row_idx}"].value)

    if cfo and not level and not kind and not comment and not project:
        return "cfo_total"
    if cfo and level and not kind and not comment and not project:
        return "level_total"
    if cfo and level:
        return "detail"
    return "other"


def _clone_templates(templates: dict[str, Any]) -> dict[str, Any]:
    return {
        key: {
            "height": value["height"],
            "styles": {c: copy(s) for c, s in value["styles"].items()},
            "merges": list(value.get("merges", [])),
        }
        for key, value in templates.items()
    }


def _clone_template(template: dict[str, Any]) -> dict[str, Any]:
    return {
        "height": template["height"],
        "styles": {c: copy(s) for c, s in template["styles"].items()},
        "merges": list(template.get("merges", [])),
    }


def _limit_template_columns(template: dict[str, Any], max_col: int) -> dict[str, Any]:
    return {
        "height": template["height"],
        "styles": {col: copy(style) for col, style in template["styles"].items() if col <= max_col},
        "merges": [
            (min_col, max_merge_col)
            for min_col, max_merge_col in template.get("merges", [])
            if max_merge_col <= max_col
        ],
    }


def _find_block(blocks: list[dict[str, Any]], cfo_name: str) -> dict[str, Any] | None:
    target = _normalize_key(cfo_name)
    for block in blocks:
        if _normalize_key(block["name"]) == target:
            return block
    return None


def _find_level(levels: list[dict[str, Any]], level_name: str) -> dict[str, Any] | None:
    target = _normalize_key(level_name)
    for level in levels:
        if _normalize_key(level["name"]) == target:
            return level
    return None


def _find_sheet_by_name(workbook: Workbook, sheet_name: str) -> Worksheet | None:
    target = _normalize_key(sheet_name)
    for candidate in workbook.sheetnames:
        if _normalize_key(candidate) == target:
            return workbook[candidate]
    return None


def _find_header_row(
    sheet: Worksheet,
    preset: dict[str, Any],
    page_name: str,
    source_name: str,
    max_scan_rows: int = 50,
) -> int:
    for row_idx in range(1, min(sheet.max_row, max_scan_rows) + 1):
        if _row_matches_header_preset(sheet, row_idx, preset):
            return row_idx
    raise ConfigError(
        f"[{page_name}] В исходном файле '{source_name}' не найдена строка заголовков на листе '{sheet.title}'."
    )


def _row_matches_header_preset(sheet: Worksheet, row_idx: int, preset: dict[str, Any]) -> bool:
    headers_in_row = {_normalize_header_name(sheet.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, sheet.max_column + 1)}
    headers_in_row.discard("")

    for spec in preset.values():
        candidates = _extract_header_candidates(spec)
        if not candidates:
            continue
        normalized_candidates = {_normalize_header_name(candidate) for candidate in candidates}
        if headers_in_row.isdisjoint(normalized_candidates):
            return False
    return True


def _extract_header_candidates(spec: Any) -> list[str]:
    if not isinstance(spec, dict):
        return []

    candidates: list[str] = []
    header = spec.get("header")
    if isinstance(header, str) and header.strip():
        candidates.append(header)

    aliases = spec.get("header_aliases")
    if isinstance(aliases, list):
        candidates.extend(alias for alias in aliases if isinstance(alias, str) and alias.strip())

    return candidates


def _pick_first_project(blocks: list[dict[str, Any]]) -> str:
    for block in blocks:
        for level in block["levels"]:
            for detail in level["details"]:
                project = _clean_text(detail.get("project"))
                if project:
                    return project
    return ""


def _append_unique(items: list[str], value: str) -> None:
    if value and value not in items:
        items.append(value)


def _single_value_or_blank(values: list[str]) -> str:
    if len(values) == 1:
        return values[0]
    return ""


def _project_value_or_blank(values: list[str]) -> str:
    if not values:
        return ""
    return values[0]


def _set_row_grouping(
    sheet: Worksheet,
    row_idx: int,
    outline_level: int,
    hidden: bool,
    collapsed: bool,
) -> None:
    row_dimension = sheet.row_dimensions[row_idx]
    row_dimension.outlineLevel = outline_level
    row_dimension.hidden = hidden
    row_dimension.collapsed = collapsed


def _extract_year(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)):
        if 1900 <= int(value) <= 2100:
            return int(value)
        return value

    text = _clean_text(value)
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 4:
        return int(digits[:4])
    return text


def _normalize_type(value: Any, type_map: Any) -> str:
    text = _clean_text(value)
    if isinstance(type_map, dict):
        mapped = type_map.get(text)
        if mapped is None and text.isdigit():
            mapped = type_map.get(str(int(text)))
        if isinstance(mapped, str) and mapped.strip():
            return mapped.strip()
    return text


def _force_full_recalculation(workbook: Workbook) -> None:
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass


def _get_preset(presets: Any, preset_name: str, page_name: str) -> dict[str, Any]:
    if not isinstance(presets, dict):
        raise ConfigError(f"[{page_name}] Неверно описаны пресеты колонок.")
    preset = presets.get(preset_name)
    if not isinstance(preset, dict):
        raise ConfigError(f"[{page_name}] Не найден пресет '{preset_name}'.")
    return preset


def _resolve_preset_column(
    sheet: Worksheet,
    preset: dict[str, Any],
    field_name: str,
    header_row: int,
    context: str,
) -> int:
    if field_name not in preset:
        raise ConfigError(f"{context}: в пресете нет поля '{field_name}'.")
    return _resolve_column(sheet, preset[field_name], header_row, f"{context}, поле '{field_name}'")


def _resolve_column(
    sheet: Worksheet,
    spec: Any,
    header_row: int,
    context: str,
) -> int:
    if isinstance(spec, int):
        return spec

    if isinstance(spec, str):
        return _parse_column_reference(spec)

    if not isinstance(spec, dict):
        raise ConfigError(f"{context}: описание колонки должно быть строкой, числом или объектом.")

    if "column" in spec:
        return _parse_column_reference(spec["column"])

    header_candidates: list[str] = []
    if isinstance(spec.get("header"), str):
        header_candidates.append(spec["header"])
    aliases = spec.get("header_aliases")
    if isinstance(aliases, list):
        header_candidates.extend(alias for alias in aliases if isinstance(alias, str))

    if not header_candidates:
        raise ConfigError(f"{context}: у колонки должен быть 'column', 'header' или 'header_aliases'.")

    normalized_headers = {_normalize_header_name(name) for name in header_candidates}
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if _normalize_header_name(cell_value) in normalized_headers:
            return col_idx

    raise ConfigError(
        f"{context}: не удалось найти колонку по заголовкам {', '.join(header_candidates)}."
    )


def _parse_column_reference(value: Any) -> int:
    if isinstance(value, int):
        return value
    if not isinstance(value, str):
        raise ConfigError("Ссылка на колонку должна быть строкой или числом.")

    candidate = value.strip()
    if not candidate:
        raise ConfigError("Пустая ссылка на колонку.")
    if candidate.isdigit():
        return int(candidate)

    try:
        return column_index_from_string(candidate.upper())
    except ValueError as exc:
        raise ConfigError(f"Некорректная ссылка на колонку: '{value}'.") from exc


def _to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0.0

    try:
        return float(text)
    except ValueError as exc:
        raise ConfigError(f"Значение '{value}' нельзя привести к числу.") from exc


def _normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return " ".join(text.lower().replace("ё", "е").split())


def _normalize_header_name(value: Any) -> str:
    return _normalize_key(value)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _reset_row_merges(sheet: Worksheet, row_idx: int) -> None:
    ranges_to_remove = [
        str(merged_range)
        for merged_range in sheet.merged_cells.ranges
        if merged_range.min_row == row_idx and merged_range.max_row == row_idx
    ]
    for range_ref in ranges_to_remove:
        sheet.unmerge_cells(range_ref)


def _set_cell_value(sheet: Worksheet, coordinate: str, value: Any) -> None:
    if isinstance(sheet[coordinate], MergedCell):
        return
    sheet[coordinate] = value


def _estimate_wrapped_lines(value: Any, column_width: float | None) -> int:
    if value in (None, ""):
        return 1

    text = str(value)
    width = max(int((column_width or 10) * 1.1), 1)
    line_count = 0
    for raw_line in text.splitlines() or [""]:
        line = raw_line or " "
        line_count += max((len(line) + width - 1) // width, 1)
    return max(line_count, 1)


def _format_cfo_row(sheet: Worksheet, row_idx: int, min_height: float | None) -> None:
    max_lines = 1
    for column in ("A", "B", "C"):
        cell = sheet[f"{column}{row_idx}"]
        alignment = copy(cell.alignment)
        alignment.vertical = "center"
        alignment.wrap_text = True
        cell.alignment = alignment

        column_letter = column
        width = sheet.column_dimensions[column_letter].width
        if column in {"A", "C"}:
            max_lines = max(max_lines, _estimate_wrapped_lines(cell.value, width))

    base_height = float(min_height) if min_height else 15.0
    sheet.row_dimensions[row_idx].height = max(base_height, 15.0 * max_lines)


def _require_str(payload: dict[str, Any], key: str, page_name: str) -> str:
    value = payload.get(key)
    if not isinstance(value, str) or not value.strip():
        raise ConfigError(f"[{page_name}] Поле '{key}' должно быть непустой строкой.")
    return value
