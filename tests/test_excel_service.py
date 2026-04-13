from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.excel_service import consolidate_workbooks


def test_preserves_row_dimensions_for_existing_and_new_reserve_rows() -> None:
    template = _build_template_workbook(include_cfo=False)
    source = _build_source_workbook()
    config = _build_config(include_cfo=False)

    result_bytes, _, report = consolidate_workbooks(
        template_bytes=_dump_book(template),
        template_name="template.xlsx",
        sources=[("source.xlsx", _dump_book(source))],
        config=config,
    )

    result_book = load_workbook(BytesIO(result_bytes))
    reserves_ws = result_book["Справка по резервам "]
    helper_ws = result_book["Справка"]

    assert report["pages"][0]["detail_rows_written"] == 3
    assert reserves_ws.row_dimensions[9].height == 16.5
    assert reserves_ws.row_dimensions[10].height == 45
    assert reserves_ws.row_dimensions[11].outlineLevel == 2
    assert reserves_ws["B10"].border.bottom.style != "medium"
    assert helper_ws.row_dimensions[10].height == 20
    assert helper_ws.row_dimensions[11].height == 45


def test_preserves_cfo_comment_visibility_from_template() -> None:
    template = _build_template_workbook(include_cfo=True)
    source = _build_source_workbook()
    config = _build_config(include_cfo=True)

    result_bytes, _, _ = consolidate_workbooks(
        template_bytes=_dump_book(template),
        template_name="template.xlsx",
        sources=[("source.xlsx", _dump_book(source))],
        config=config,
    )

    result_book = load_workbook(BytesIO(result_bytes))
    cfo_ws = result_book["Справка по ЦФО"]

    assert cfo_ws.row_dimensions[4].hidden is False
    assert cfo_ws.row_dimensions[4].collapsed is False
    assert cfo_ws["B4"].alignment.wrap_text is None
    assert cfo_ws.row_dimensions[5].outlineLevel == 1
    assert cfo_ws.row_dimensions[5].hidden is False
    assert cfo_ws.row_dimensions[5].collapsed is True
    assert cfo_ws.row_dimensions[6].outlineLevel == 2
    assert cfo_ws.row_dimensions[6].hidden is True


def test_new_cfo_sheet_rows_do_not_inherit_template_fill() -> None:
    template = _build_template_workbook(include_cfo=True)
    source = _build_source_workbook(new_cfo=True)
    config = _build_config(include_cfo=True)

    result_bytes, _, _ = consolidate_workbooks(
        template_bytes=_dump_book(template),
        template_name="template.xlsx",
        sources=[("source.xlsx", _dump_book(source))],
        config=config,
    )

    result_book = load_workbook(BytesIO(result_bytes))
    cfo_ws = result_book["Справка по ЦФО"]

    assert cfo_ws["A8"].fill.fill_type == "solid"
    assert cfo_ws["B8"].fill.fill_type is None
    assert cfo_ws["C8"].fill.fill_type == "solid"


def test_cfo_sheet_clears_fill_for_cfo_new_in_reserve_layout() -> None:
    template = _build_template_workbook(include_cfo=True, cfo_sheet_name="ЦФО 2")
    source = _build_source_workbook(new_cfo=True, new_cfo_name="ЦФО 2")
    config = _build_config(include_cfo=True)

    result_bytes, _, _ = consolidate_workbooks(
        template_bytes=_dump_book(template),
        template_name="template.xlsx",
        sources=[("source.xlsx", _dump_book(source))],
        config=config,
    )

    result_book = load_workbook(BytesIO(result_bytes))
    cfo_ws = result_book["Справка по ЦФО"]

    assert cfo_ws["A4"].value == "Вскрытие ЦФО 2"
    assert cfo_ws["A4"].fill.fill_type == "solid"
    assert cfo_ws["B4"].fill.fill_type is None
    assert cfo_ws["C4"].fill.fill_type == "solid"


def test_new_cfo_block_uses_same_pastel_fill_for_cfo_cells() -> None:
    template = _build_template_workbook(include_cfo=False)
    source = _build_source_workbook(new_cfo=True)
    config = _build_config(include_cfo=False)

    result_bytes, _, _ = consolidate_workbooks(
        template_bytes=_dump_book(template),
        template_name="template.xlsx",
        sources=[("source.xlsx", _dump_book(source))],
        config=config,
    )

    result_book = load_workbook(BytesIO(result_bytes))
    reserves_ws = result_book["Справка по резервам "]
    expected_color = "00E2F0D9"

    for cell_ref in ("B12", "B13", "B14"):
        assert reserves_ws[cell_ref].fill.fill_type == "solid"
        assert reserves_ws[cell_ref].fill.start_color.rgb == expected_color
    assert reserves_ws["E12"].fill.fill_type == "solid"
    assert reserves_ws["E12"].fill.start_color.rgb == expected_color
    assert reserves_ws["E13"].fill.fill_type is None


def _build_template_workbook(*, include_cfo: bool, cfo_sheet_name: str = "ЦФО 1") -> Workbook:
    workbook = Workbook()
    reserves_ws = workbook.active
    reserves_ws.title = "Справка по резервам "

    reserves_ws["B7"] = "ЦФО 1"
    reserves_ws["F7"] = "=F8"
    reserves_ws["B8"] = "ЦФО 1"
    reserves_ws["D8"] = "РП"
    reserves_ws["F8"] = "=SUM(F9:F10)"
    reserves_ws["B9"] = "ЦФО 1"
    reserves_ws["D9"] = "РП"
    reserves_ws["E9"] = "Работа 1"
    reserves_ws["F9"] = 10
    reserves_ws["G9"] = "Удорожание"
    reserves_ws["H9"] = "Комментарий 1"
    reserves_ws["I9"] = 2024
    reserves_ws["J9"] = "Проект 1"
    reserves_ws["B10"] = "ЦФО 1"
    reserves_ws["D10"] = "РП"
    reserves_ws["E10"] = "Работа 2"
    reserves_ws["F10"] = 20
    reserves_ws["G10"] = "Идеологическое изменение"
    reserves_ws["H10"] = "Очень длинный комментарий, который должен сохранить высоту строки"
    reserves_ws["I10"] = 2025
    reserves_ws["J10"] = "Проект 2"

    reserves_ws.row_dimensions[7].outlineLevel = 0
    reserves_ws.row_dimensions[8].outlineLevel = 1
    reserves_ws.row_dimensions[9].outlineLevel = 2
    reserves_ws.row_dimensions[10].outlineLevel = 2
    reserves_ws.row_dimensions[9].height = 16.5
    reserves_ws.row_dimensions[10].height = 45
    bright_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    for cell_ref in ("D7", "E7", "F7", "G7", "H7", "I7", "J7"):
        reserves_ws[cell_ref].fill = bright_fill

    helper_ws = workbook.create_sheet("Справка")
    helper_ws["B10"] = "='Справка по резервам '!D9"
    helper_ws["C10"] = "='Справка по резервам '!E9"
    helper_ws["D10"] = "='Справка по резервам '!F9"
    helper_ws["E10"] = "='Справка по резервам '!G9"
    helper_ws["F10"] = "='Справка по резервам '!H9"
    helper_ws["G10"] = "='Справка по резервам '!I9"
    helper_ws["H10"] = "='Справка по резервам '!J9"
    helper_ws["I10"] = "='Справка по резервам '!B9"
    helper_ws["B11"] = "='Справка по резервам '!D10"
    helper_ws["C11"] = "='Справка по резервам '!E10"
    helper_ws["D11"] = "='Справка по резервам '!F10"
    helper_ws["E11"] = "='Справка по резервам '!G10"
    helper_ws["F11"] = "='Справка по резервам '!H10"
    helper_ws["G11"] = "='Справка по резервам '!I10"
    helper_ws["H11"] = "='Справка по резервам '!J10"
    helper_ws["I11"] = "='Справка по резервам '!B10"
    helper_ws.row_dimensions[10].height = 20
    helper_ws.row_dimensions[11].height = 45
    helper_ws.row_dimensions[10].hidden = True
    helper_ws.row_dimensions[11].hidden = True

    if include_cfo:
        cfo_ws = workbook.create_sheet("Справка по ЦФО")
        cfo_ws["A4"] = f"Вскрытие {cfo_sheet_name}"
        cfo_ws["B4"] = 30
        cfo_ws["A5"] = "Комментарий 1"
        cfo_ws["B5"] = 30
        cfo_ws["A6"] = "Причина 1"
        cfo_ws["B6"] = 30
        cfo_ws["C6"] = "Проект 1"
        cfo_ws["A7"] = "Общий итог"
        cfo_ws["B7"] = 30

        cfo_ws.row_dimensions[4].outlineLevel = 0
        cfo_ws.row_dimensions[4].hidden = False
        cfo_ws.row_dimensions[4].collapsed = False
        cfo_ws.row_dimensions[5].outlineLevel = 1
        cfo_ws.row_dimensions[5].hidden = False
        cfo_ws.row_dimensions[5].collapsed = True
        cfo_ws.row_dimensions[6].outlineLevel = 2
        cfo_ws.row_dimensions[6].hidden = True
        cfo_ws.row_dimensions[6].collapsed = False
        cfo_ws.row_dimensions[7].outlineLevel = 0
        cfo_fill = PatternFill(fill_type="solid", start_color="CC99FF", end_color="CC99FF")
        for cell_ref in ("A4", "C4", "A5", "C5", "A6"):
            cfo_ws[cell_ref].fill = cfo_fill

    return workbook


def _build_source_workbook(*, new_cfo: bool = False, new_cfo_name: str = "ЦФО 2") -> Workbook:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Лист1"
    ws.append(
        [
            "Проект",
            "ЦФО",
            "Дата вскрытия",
            "Сумма вскрытия",
            "Причина",
            "Тип вскрытия",
            "Уровень резерва",
            "Комментарий экономиста",
        ]
    )
    ws.append(["Проект 3", "ЦФО 1", 2026, 15, "Причина 1", "Удорожание", "РП", "Комментарий 1"])
    if new_cfo:
        ws.append(["Проект 4", new_cfo_name, 2026, 25, "Причина 2", "Удорожание", "РП", "Комментарий 2"])
    return workbook


def _build_config(*, include_cfo: bool) -> dict[str, object]:
    page: dict[str, object] = {
        "name": "Справка по резервам",
        "mode": "reserves_sheet_v1",
        "template_sheet": "Справка по резервам ",
        "helper_sheet": "Справка",
        "helper_start_row": 10,
        "template_start_row": 7,
        "source_sheet": "Лист1",
        "source_preset": "default",
        "source_header_row": 1,
        "source_start_row": 2,
    }
    if include_cfo:
        page["cfo_sheet"] = "Справка по ЦФО"
        page["cfo_start_row"] = 4
        page["cfo_group_template_row"] = 4
        page["cfo_comment_template_row"] = 5
        page["cfo_reason_template_row"] = 6
        page["cfo_total_template_row"] = 7
        page["cfo_total_label"] = "Общий итог"

    return {
        "source_cols_preset": {
            "default": {
                "project": {"header": "Проект"},
                "cfo": {"header": "ЦФО"},
                "date": {"header": "Дата вскрытия"},
                "amount": {"header": "Сумма вскрытия"},
                "reason": {"header": "Причина"},
                "type": {"header": "Тип вскрытия"},
                "level": {"header": "Уровень резерва"},
                "comment": {"header": "Комментарий экономиста"},
            }
        },
        "pages": [page],
    }


def _dump_book(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
