from __future__ import annotations

from openpyxl import Workbook

from app.reserve_header_sync import (
    apply_snapshots_to_template,
    enrich_snapshots_labels,
    extract_approved_reserve_snapshots,
    find_template_reserve_plan_rows,
)


def test_applies_approved_reserves_by_project_and_appends_formulas() -> None:
    source = Workbook()
    src_ws = source.active
    src_ws.title = "Вскрытие резервов"
    src_ws.merge_cells("C4:C5")
    src_ws["C4"] = "Наименование проекта"
    src_ws.merge_cells("M4:P4")
    src_ws["M4"] = "Утвержденные резервы"
    src_ws["M5"] = "РП"
    src_ws["N5"] = "РГП"
    src_ws["O5"] = "ЗГД"
    src_ws["P5"] = "ГД"
    src_ws["C7"] = 1
    src_ws["M7"] = 10
    src_ws["N7"] = 20
    src_ws["C8"] = 1
    src_ws["M8"] = 11
    src_ws["N8"] = 21
    src_ws["C10"] = "Проект"
    src_ws["D10"] = "ЦФО"
    src_ws["E10"] = "Уровень резерва"
    src_ws["F10"] = "Дата вскрытия"
    src_ws["G10"] = "Сумма вскрытия"

    preset = {
        "project": {"header": "Проект"},
        "cfo": {"header": "ЦФО"},
        "level": {"header": "Уровень резерва"},
        "date": {"header": "Дата вскрытия"},
        "amount": {"header": "Сумма вскрытия"},
    }
    snapshots = extract_approved_reserve_snapshots(src_ws, preset=preset)
    enrich_snapshots_labels(snapshots, [{"project": "1"}])

    template = Workbook()
    tpl_ws = template.active
    tpl_ws.title = "Справка по резервам "
    tpl_ws["B3"] = "Наименование проекта"
    tpl_ws["B4"] = "Проект 1"
    tpl_ws["D5"] = "План по ПД с резервами"
    tpl_ws["E5"] = "РП"
    tpl_ws["F5"] = "РГП"
    tpl_ws["G5"] = "ЗГД"
    tpl_ws["H5"] = "ГД"
    tpl_ws["E6"] = 100
    tpl_ws["F6"] = "=50+25"

    apply_snapshots_to_template(tpl_ws, snapshots, find_template_reserve_plan_rows(tpl_ws))

    assert tpl_ws["E6"].value == "=100+10+11"
    assert tpl_ws["F6"].value == "=50+25+20+21"
