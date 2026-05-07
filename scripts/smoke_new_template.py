from pathlib import Path

from app.config import load_app_config
from app.excel_service import consolidate_workbooks

ROOT = Path(__file__).resolve().parents[1]


def main() -> None:
    from openpyxl import load_workbook

    cfg = load_app_config()
    cfg["default_template_path"] = "Файлы/new/Шаблон.xlsm"
    # На шаблоне «Файлы/new/Шаблон.xlsm» первая строка блока ЦФО — 9 (итог ЦФО первой таблицы).
    cfg["pages"] = [
        {
            **page,
            "template_start_row": 9,
            "preserve_reserve_sector_gap": {
                "enabled": True,
                "source_row_first": 33,
                "source_row_last": 37,
            },
            "reserve_header_sync": {
                **(page.get("reserve_header_sync") if isinstance(page.get("reserve_header_sync"), dict) else {}),
                "enabled": True,
            },
        }
        for page in cfg.get("pages", [])
    ]

    src = ROOT / "Файлы" / "new" / "1.xlsx"
    b, name, rep = consolidate_workbooks(
        None,
        None,
        [(src.name, src.read_bytes())],
        cfg,
    )
    out = ROOT / "Файлы" / "new" / "_smoke_result.xlsm"
    out.write_bytes(b)

    s = load_workbook(out, data_only=True)["Справка по резервам "]
    print(name, rep["pages"][0], "->", out)
    print(
        "Cells:",
        "B4",
        s["B4"].value,
        "B34",
        s["B34"].value,
        "E6:H6",
        [s.cell(6, c).value for c in range(5, 9)],
        "E36:H36",
        [s.cell(36, c).value for c in range(5, 9)],
    )


if __name__ == "__main__":
    main()
